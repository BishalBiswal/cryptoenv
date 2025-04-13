from pycoingecko import CoinGeckoAPI # type: ignore
from openpyxl import Workbook,load_workbook# type: ignore
crypto_list=['bitcoin','ethereum']

def crypto_prices(cryptos):
    cg=CoinGeckoAPI()
    response=cg.get_price(ids=','.join(cryptos),vs_currencies = 'usd')
    return response

prices=crypto_prices(crypto_list)

def update_excel():
    try:
        workbook=load_workbook('Crypto.xlsx')
        sheet=workbook.active
    except FileNotFoundError:
        workbook=Workbook()
        sheet=workbook.active
        sheet['A1']='Crypto Name'
        sheet['B1']='Price'

    existing_crypto=[]
    for row in sheet.iter_rows(min_row=2,max_col=1,values_only=2):
        existing_crypto.append(row[0])

    for cryptos in crypto_list:
        price=prices[cryptos]['usd']
        if cryptos in existing_crypto:
            for row in sheet.iter_rows(min_row=2,max_column=2):
                if row[0].value==cryptos:
                    row[1].value= price
                    break
        else:
            sheet.append([cryptos, price])

    workbook.save('Crypto.xlsx')
update_excel()